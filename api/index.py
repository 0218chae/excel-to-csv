# api/index.py (Vercel 서버리스 함수용 Flask + openpyxl 경량 구현)
import io, os, re, csv, zipfile
from flask import Flask, request, send_file, abort
from openpyxl import load_workbook

app = Flask(__name__)

# 한글/영문/숫자/밑줄/하이픈/공백 허용
_invalid = r"[^\w\-\u3131-\u318E\uAC00-\uD7A3\s]"
_space_re = re.compile(r"\s+")

def safe_name(name: str) -> str:
    name = re.sub(_invalid, "_", name)
    return _space_re.sub(" ", name).strip() or "sheet"

# 파일 크기 제한 (예: 12MB) — 필요하면 조절
MAX_BYTES = 12 * 1024 * 1024


def _convert_impl(file_storage):
    if not file_storage.filename:
        abort(400, "파일 이름이 비어 있습니다.")
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in [".xlsx", ".xlsm"]:  # openpyxl은 .xlsx/.xlsm 권장
        abort(400, "xlsx/xlsm만 허용합니다. (구형 .xls는 지원하지 않음)")

    raw = file_storage.read()
    if len(raw) > MAX_BYTES:
        abort(413, f"파일이 너무 큽니다. 최대 {MAX_BYTES // (1024*1024)}MB")

    buf = io.BytesIO(raw)
    buf.seek(0)

    try:
        wb = load_workbook(buf, data_only=True, read_only=True)
    except Exception as e:
        abort(400, f"엑셀 읽기 오류: {e}")

    if not wb.sheetnames:
        abort(400, "시트가 없습니다.")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        used = {}
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                out = io.StringIO()
                writer = csv.writer(out, lineterminator="\n")
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([("" if v is None else v) for v in row])
                csv_bytes = out.getvalue().encode("utf-8-sig")
            except Exception as e:
                zf.writestr(f"ERROR_{safe_name(sheet_name)}.txt", str(e))
                continue

            base = safe_name(sheet_name)
            n = used.get(base, 0)
            used[base] = n + 1
            name = f"{base}.csv" if n == 0 else f"{base}_{n+1}.csv"
            zf.writestr(name, csv_bytes)

    zip_buf.seek(0)
    base = safe_name(os.path.splitext(file_storage.filename)[0]) or "excel"
    return send_file(
        zip_buf,
        as_attachment=True,
        download_name=f"{base}_sheets.zip",
        mimetype="application/zip",
    )


@app.post("/api/convert")
def convert_route():
    if "file" not in request.files:
        abort(400, "업로드된 파일이 없습니다.")
    return _convert_impl(request.files["file"])


# Vercel에서 파일명이 index.py인 경우 /api/index 로 접근하는 사용자를 위해 추가 라우트 지원
@app.post("/api/index")
def index_route():
    if "file" not in request.files:
        abort(400, "업로드된 파일이 없습니다.")
    return _convert_impl(request.files["file"])